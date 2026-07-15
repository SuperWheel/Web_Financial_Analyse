"""Excel 模板生成与三表导入（忽略财务比率 sheet）。"""
from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.constants import (
    BALANCE_SHEET_FIELDS,
    BALANCE_SHEET_GROUPS,
    CASH_FLOW_FIELDS,
    CASH_FLOW_GROUPS,
    INCOME_STATEMENT_FIELDS,
    INCOME_STATEMENT_GROUPS,
    PERIOD_ANNUAL,
    PERIOD_QUARTERLY,
)
from app.models.balance_sheet import BalanceSheet
from app.models.cash_flow import CashFlowStatement
from app.models.income_statement import IncomeStatement
from app.schemas.statement import (
    BalanceSheetCreate,
    BalanceSheetUpdate,
    CashFlowStatementCreate,
    CashFlowStatementUpdate,
    IncomeStatementCreate,
    IncomeStatementUpdate,
)
from app.services.company_service import get_company
from app.services.exceptions import ServiceError
from app.services.statement_service import _find_by_period

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_GROUP_FILL = PatternFill("solid", fgColor="D6EAF8")
_GROUP_FONT = Font(bold=True)

_SHEET_MAP: dict[str, dict[str, Any]] = {
    "资产负债表": {
        "statement_type": "balance",
        "groups": BALANCE_SHEET_GROUPS,
        "fields": set(BALANCE_SHEET_FIELDS),
        "model": BalanceSheet,
        "create_schema": BalanceSheetCreate,
        "update_schema": BalanceSheetUpdate,
    },
    "利润表": {
        "statement_type": "income",
        "groups": INCOME_STATEMENT_GROUPS,
        "fields": set(INCOME_STATEMENT_FIELDS),
        "model": IncomeStatement,
        "create_schema": IncomeStatementCreate,
        "update_schema": IncomeStatementUpdate,
    },
    "现金流量表": {
        "statement_type": "cashflow",
        "groups": CASH_FLOW_GROUPS,
        "fields": set(CASH_FLOW_FIELDS),
        "model": CashFlowStatement,
        "create_schema": CashFlowStatementCreate,
        "update_schema": CashFlowStatementUpdate,
    },
}

# 允许英文别名 sheet
_SHEET_ALIASES = {
    "balance": "资产负债表",
    "balance_sheet": "资产负债表",
    "income": "利润表",
    "income_statement": "利润表",
    "cashflow": "现金流量表",
    "cash_flow": "现金流量表",
    "cash_flow_statement": "现金流量表",
}


class ValidationError(ServiceError):
    status_code = 422


_ANNUAL_RE = re.compile(r"^(\d{4})\s*年报$")
_QUARTER_RE = re.compile(r"^(\d{4})\s*[Qq]([1-4])$")


def parse_period_header(text: str) -> dict[str, Any] | None:
    if text is None:
        return None
    s = str(text).strip()
    if not s:
        return None
    m = _ANNUAL_RE.match(s)
    if m:
        y = int(m.group(1))
        return {
            "year": y,
            "period_type": PERIOD_ANNUAL,
            "quarter": None,
            "label": f"{y} 年报",
        }
    m = _QUARTER_RE.match(s)
    if m:
        y = int(m.group(1))
        q = int(m.group(2))
        return {
            "year": y,
            "period_type": PERIOD_QUARTERLY,
            "quarter": q,
            "label": f"{y} Q{q}",
        }
    return None


def _period_label(year: int, period_type: str, quarter: int | None) -> str:
    if period_type == PERIOD_ANNUAL:
        return f"{year} 年报"
    return f"{year} Q{quarter}"


def _style_header(ws, n: int) -> None:
    for col in range(1, n + 1):
        cell = ws.cell(1, col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize(ws, max_width: int = 28) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def build_template_workbook(
    *,
    period_type: str = PERIOD_ANNUAL,
    years: list[int] | None = None,
    quarters: list[int] | None = None,
) -> tuple[bytes, str]:
    """空模板：三表 + 说明。"""
    if period_type not in (PERIOD_ANNUAL, PERIOD_QUARTERLY):
        raise ValidationError("period_type 须为 annual 或 quarterly")

    if not years:
        # 默认近 3 个日历年（占位，用户可改列头）
        y = datetime.now().year
        years = [y - 2, y - 1, y]

    periods: list[dict[str, Any]] = []
    if period_type == PERIOD_ANNUAL:
        for y in sorted(years):
            periods.append(
                {
                    "year": y,
                    "period_type": PERIOD_ANNUAL,
                    "quarter": None,
                    "label": _period_label(y, PERIOD_ANNUAL, None),
                }
            )
    else:
        qs = quarters or [1, 2, 3, 4]
        for y in sorted(years):
            for q in qs:
                periods.append(
                    {
                        "year": y,
                        "period_type": PERIOD_QUARTERLY,
                        "quarter": q,
                        "label": _period_label(y, PERIOD_QUARTERLY, q),
                    }
                )

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws0 = wb.create_sheet("说明", 0)
    ws0.append(["项", "内容"])
    _style_header(ws0, 2)
    for k, v in [
        ("用途", "Excel 模板导入 — 填写三表金额后上传"),
        ("期间类型", "年报" if period_type == PERIOD_ANNUAL else "季报"),
        ("期间列", "、".join(p["label"] for p in periods)),
        ("科目代码", "请勿改「科目代码」列；可改金额"),
        ("财务比率", "勿在本模板填写比率；入库后系统动态计算"),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]:
        ws0.append([k, v])
    _autosize(ws0, 50)

    for title, cfg in _SHEET_MAP.items():
        ws = wb.create_sheet(title)
        headers = ["科目代码", "科目"] + [p["label"] for p in periods]
        ws.append(headers)
        _style_header(ws, len(headers))
        for g in cfg["groups"]:
            ws.append(["", g["label"]] + [None] * len(periods))
            gr = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(gr, col)
                cell.fill = _GROUP_FILL
                cell.font = _GROUP_FONT
            for f in g["fields"]:
                ws.append([f["key"], f["label"]] + [None] * len(periods))
        _autosize(ws)
        ws.freeze_panes = "C2"

    # 提示 sheet：比率不导入
    ws_r = wb.create_sheet("财务比率")
    ws_r.append(["说明"])
    ws_r.append(["本 sheet 仅作占位/导出兼容；导入时忽略。比率由系统按科目动态计算。"])
    _autosize(ws_r, 60)

    buf = BytesIO()
    wb.save(buf)
    years_span = f"{min(years)}-{max(years)}" if years else "template"
    fname = f"财务三表模板_{period_type}_{years_span}.xlsx"
    return buf.getvalue(), fname


def _cell_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip().replace(",", "")
        if not s or s in {"—", "-", "–"}:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(val, bool):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _resolve_sheet_name(name: str) -> str | None:
    if name in _SHEET_MAP:
        return name
    key = name.strip().lower().replace(" ", "_")
    return _SHEET_ALIASES.get(key)


def _parse_workbook(content: bytes) -> tuple[str | None, list[dict[str, Any]], list[str]]:
    """
    返回 (period_type, sheet_payloads, warnings)
    sheet_payload: {
      statement_type, label, periods: [...],
      # period_key -> {field: value}
      data: dict[tuple, dict[str, float|None]]
    }
    """
    warnings: list[str] = []
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValidationError(f"无法读取 Excel：{exc}") from exc

    sheet_payloads: list[dict[str, Any]] = []
    workbook_period_type: str | None = None

    for sheet_name in wb.sheetnames:
        resolved = _resolve_sheet_name(sheet_name)
        if resolved is None:
            if sheet_name not in ("说明", "财务比率"):
                warnings.append(f"忽略未知工作表：{sheet_name}")
            continue
        cfg = _SHEET_MAP[resolved]
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            warnings.append(f"{resolved} 为空表")
            continue

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        if len(header) < 3 or header[0] not in ("科目代码", "代码", "key"):
            warnings.append(f"{resolved} 表头须以「科目代码」开头，已跳过")
            continue

        period_cols: list[tuple[int, dict[str, Any]]] = []
        for idx, h in enumerate(header[2:], start=2):
            if not h:
                continue
            p = parse_period_header(h)
            if p is None:
                warnings.append(f"{resolved} 无法识别期间列「{h}」，已忽略")
                continue
            if workbook_period_type is None:
                workbook_period_type = p["period_type"]
            elif p["period_type"] != workbook_period_type:
                raise ValidationError(
                    f"工作簿期间类型混用：已有 {workbook_period_type}，列「{h}」为 {p['period_type']}"
                )
            period_cols.append((idx, p))

        if not period_cols:
            warnings.append(f"{resolved} 无有效期间列")
            continue

        # data[period_tuple][field] = value
        data: dict[tuple[int, str, int | None], dict[str, float | None]] = {
            (p["year"], p["period_type"], p["quarter"]): {} for _, p in period_cols
        }
        known = cfg["fields"]
        rows_with_code = 0
        non_null = 0

        for row in rows[1:]:
            if not row:
                continue
            code = row[0]
            if code is None or str(code).strip() == "":
                continue
            code_s = str(code).strip()
            rows_with_code += 1
            if code_s not in known:
                warnings.append(f"{resolved} 未知科目代码：{code_s}")
                continue
            for col_idx, p in period_cols:
                if col_idx >= len(row):
                    continue
                val = _cell_float(row[col_idx])
                key = (p["year"], p["period_type"], p["quarter"])
                data[key][code_s] = val
                if val is not None:
                    non_null += 1

        periods_meta = [p for _, p in period_cols]
        # 去重保持顺序
        seen: set[tuple[int, str, int | None]] = set()
        uniq_periods: list[dict[str, Any]] = []
        for p in periods_meta:
            k = (p["year"], p["period_type"], p["quarter"])
            if k in seen:
                continue
            seen.add(k)
            uniq_periods.append(p)

        sheet_payloads.append(
            {
                "statement_type": cfg["statement_type"],
                "label": resolved,
                "cfg": cfg,
                "periods": uniq_periods,
                "data": data,
                "non_null_fields": non_null,
                "rows_with_code": rows_with_code,
            }
        )

    if not sheet_payloads:
        raise ValidationError("未解析到任何有效三表数据（请检查 sheet 名与表头）")

    return workbook_period_type, sheet_payloads, warnings


def _period_tag(p: dict[str, Any], sheet_label: str) -> str:
    return f"{sheet_label}/{p['label']}"


def preview_excel_import(
    db: Session, company_id: int, content: bytes
) -> dict[str, Any]:
    company = get_company(db, company_id)
    period_type, sheets, warnings = _parse_workbook(content)

    # 合并期间
    all_periods: dict[tuple[int, str, int | None], dict[str, Any]] = {}
    for s in sheets:
        for p in s["periods"]:
            k = (p["year"], p["period_type"], p["quarter"])
            all_periods[k] = p

    periods_sorted = sorted(
        all_periods.values(), key=lambda p: (p["year"], p["quarter"] or 0)
    )

    will_create: list[str] = []
    will_update: list[str] = []
    will_skip: list[str] = []

    for s in sheets:
        cfg = s["cfg"]
        model = cfg["model"]
        for p in s["periods"]:
            key = (p["year"], p["period_type"], p["quarter"])
            fields = s["data"].get(key) or {}
            tag = _period_tag(p, s["label"])
            if not any(v is not None for v in fields.values()):
                will_skip.append(tag)
                continue
            existing = _find_by_period(
                db,
                model,
                company.id,
                p["year"],
                p["period_type"],
                p["quarter"],
            )
            if existing is None:
                will_create.append(tag)
            else:
                will_update.append(tag)

    return {
        "company_id": company.id,
        "period_type": period_type,
        "periods": periods_sorted,
        "sheets": [
            {
                "statement_type": s["statement_type"],
                "label": s["label"],
                "periods": s["periods"],
                "non_null_fields": s["non_null_fields"],
                "rows_with_code": s["rows_with_code"],
            }
            for s in sheets
        ],
        "warnings": warnings,
        "will_create": will_create,
        "will_update": will_update,
        "will_skip_empty": will_skip,
    }


def commit_excel_import(
    db: Session,
    company_id: int,
    content: bytes,
    *,
    overwrite: bool = True,
) -> dict[str, Any]:
    company = get_company(db, company_id)
    _period_type, sheets, warnings = _parse_workbook(content)

    created: list[str] = []
    updated: list[str] = []
    skipped: list[str] = []
    statement_ids: dict[str, Any] = {}

    # 收集写操作，最后统一 commit（create/update 内部会 commit — 需改用无 commit 路径）
    # 现有 statement_service 每次 create/update 都 commit。为简单一致，沿用逐条 commit；
    # 失败时抛错由调用方处理（已写入部分可能残留 — 文档说明）。
    # 更稳：直接操作 ORM 后单次 commit。
    try:
        for s in sheets:
            cfg = s["cfg"]
            model = cfg["model"]
            for p in s["periods"]:
                key = (p["year"], p["period_type"], p["quarter"])
                fields = s["data"].get(key) or {}
                tag = _period_tag(p, s["label"])
                payload_fields = {k: v for k, v in fields.items() if v is not None}
                # 仍写入显式 null？仅非 null 字段；全空 skip
                if not payload_fields and not any(v is not None for v in fields.values()):
                    skipped.append(tag + "（空）")
                    continue
                # 允许全 null 但有 key？上面已 skip
                if not any(v is not None for v in fields.values()):
                    skipped.append(tag + "（空）")
                    continue

                existing = _find_by_period(
                    db,
                    model,
                    company.id,
                    p["year"],
                    p["period_type"],
                    p["quarter"],
                )
                body = {
                    "year": p["year"],
                    "period_type": p["period_type"],
                    "quarter": p["quarter"],
                    **{k: fields.get(k) for k in cfg["fields"]},
                }

                if existing is None:
                    schema = cfg["create_schema"]
                    row = schema.model_validate(body)
                    # 直写 ORM，避免嵌套 commit
                    obj = model(company_id=company.id, **row.model_dump())  # type: ignore[call-arg]
                    db.add(obj)
                    db.flush()
                    created.append(tag)
                    statement_ids[tag] = obj.id
                else:
                    if not overwrite:
                        skipped.append(tag + "（已存在，overwrite=false）")
                        continue
                    # 只更新科目金额，不触碰 year/period_type
                    for f in cfg["fields"]:
                        setattr(existing, f, fields.get(f))
                    if p["period_type"] == PERIOD_ANNUAL:
                        existing.quarter = None  # type: ignore[attr-defined]
                    db.flush()
                    updated.append(tag)
                    statement_ids[tag] = existing.id

        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "company_id": company.id,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "warnings": warnings,
        "statement_ids": statement_ids,
    }
