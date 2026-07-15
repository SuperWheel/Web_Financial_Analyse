"""Excel 导出：三表科目矩阵 + 财务比率，openpyxl 生成，不落库。"""
from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.constants import (
    BALANCE_SHEET_GROUPS,
    CASH_FLOW_GROUPS,
    INCOME_STATEMENT_GROUPS,
    PERIOD_ANNUAL,
    PERIOD_QUARTERLY,
    RATIO_DEFINITIONS,
)
from app.services.company_service import get_company
from app.services.exceptions import ServiceError
from app.services.ratio_service import compute_period_ratios
from app.services.statement_service import (
    list_balance_sheets,
    list_cash_flow_statements,
    list_income_statements,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_GROUP_FILL = PatternFill("solid", fgColor="D6EAF8")
_GROUP_FONT = Font(bold=True)
_MONEY_FORMAT = "#,##0.00"
_RATIO_FORMAT = "0.0000"
_PCT_FORMAT = "0.00"


class ValidationError(ServiceError):
    status_code = 422


def _period_label(year: int, period_type: str, quarter: int | None) -> str:
    if period_type == PERIOD_ANNUAL:
        return f"{year} 年报"
    return f"{year} Q{quarter}"


def _period_key(year: int, period_type: str, quarter: int | None) -> tuple[int, str, int | None]:
    return (int(year), period_type, quarter if quarter is not None else None)


def _money(row: Any | None, field: str) -> float | None:
    if row is None:
        return None
    val = getattr(row, field, None)
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _safe_filename(name: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", name).strip() or "export"
    return cleaned[:80]


def _collect_period_axis(
    rows_list: list[list[Any]],
    period_type: str,
    years: list[int] | None,
) -> list[dict[str, Any]]:
    """三表并集期间，升序。"""
    year_set = set(years) if years else None
    seen: dict[tuple[int, str, int | None], dict[str, Any]] = {}
    for rows in rows_list:
        for r in rows:
            if r.period_type != period_type:
                continue
            y = int(r.year)
            if year_set is not None and y not in year_set:
                continue
            q = r.quarter if r.quarter is not None else None
            key = _period_key(y, r.period_type, q)
            if key not in seen:
                seen[key] = {
                    "year": y,
                    "period_type": r.period_type,
                    "quarter": q,
                    "label": _period_label(y, r.period_type, q),
                }
    periods = list(seen.values())
    periods.sort(key=lambda p: (p["year"], p["quarter"] or 0))
    return periods


def _index_by_period(rows: list[Any], period_type: str) -> dict[tuple[int, str, int | None], Any]:
    out: dict[tuple[int, str, int | None], Any] = {}
    for r in rows:
        if r.period_type != period_type:
            continue
        q = r.quarter if r.quarter is not None else None
        out[_period_key(int(r.year), r.period_type, q)] = r
    return out


def _style_header(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(1, col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize(ws, max_width: int = 28) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _write_statement_sheet(
    wb: Workbook,
    title: str,
    groups: list[dict[str, Any]],
    periods: list[dict[str, Any]],
    row_index: dict[tuple[int, str, int | None], Any],
) -> None:
    ws = wb.create_sheet(title)
    headers = ["科目代码", "科目"] + [p["label"] for p in periods]
    ws.append(headers)
    _style_header(ws, len(headers))

    for g in groups:
        # 分组标题行
        ws.append(["", g["label"]] + [None] * len(periods))
        group_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(group_row, col)
            cell.fill = _GROUP_FILL
            cell.font = _GROUP_FONT

        for f in g["fields"]:
            key = f["key"]
            label = f["label"]
            values: list[Any] = []
            for p in periods:
                pk = _period_key(p["year"], p["period_type"], p["quarter"])
                values.append(_money(row_index.get(pk), key))
            ws.append([key, label, *values])
            r = ws.max_row
            for i in range(len(periods)):
                cell = ws.cell(r, 3 + i)
                if cell.value is not None:
                    cell.number_format = _MONEY_FORMAT

    _autosize(ws)
    ws.freeze_panes = "C2"


def _write_ratio_sheet(
    wb: Workbook,
    db: Session,
    company_id: int,
    periods: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("财务比率")
    headers = ["分组", "比率", "单位", "公式"] + [p["label"] for p in periods]
    ws.append(headers)
    _style_header(ws, len(headers))

    # 预计算各期
    snapshots: list[dict[str, dict[str, Any]]] = []
    for p in periods:
        snap = compute_period_ratios(
            db,
            company_id,
            year=p["year"],
            period_type=p["period_type"],
            quarter=p["quarter"],
        )
        by_key = {item["key"]: item for item in snap["ratios"]}
        snapshots.append(by_key)

    for meta in RATIO_DEFINITIONS:
        key = meta["key"]
        unit = meta["unit"]
        unit_label = "%" if unit == "percent" else "倍"
        row_vals: list[Any] = [meta["group"], meta["name"], unit_label, meta["description"]]
        for by_key in snapshots:
            item = by_key.get(key) or {}
            raw = item.get("value")
            if raw is None:
                row_vals.append(None)
            elif unit == "percent":
                row_vals.append(round(float(raw) * 100.0, 4))
            else:
                row_vals.append(round(float(raw), 6))
        ws.append(row_vals)
        r = ws.max_row
        for i in range(len(periods)):
            cell = ws.cell(r, 5 + i)
            if cell.value is None:
                continue
            cell.number_format = _PCT_FORMAT if unit == "percent" else _RATIO_FORMAT

    # 附注
    ws.append([])
    note_row = ws.max_row + 1
    ws.cell(
        note_row,
        1,
        "说明：单位为「%」的列为百分数（如 12.5 表示 12.5%）；「倍」为倍数原值。缺数留空。",
    )
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=max(4, 4 + len(periods)))

    _autosize(ws, max_width=36)
    ws.freeze_panes = "E2"


def _write_readme(
    wb: Workbook,
    company_name: str,
    company_code: str | None,
    period_type: str,
    periods: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("说明", 0)
    lines = [
        ("企业", company_name),
        ("代码", company_code or "—"),
        ("期间类型", "年报" if period_type == PERIOD_ANNUAL else "季报"),
        ("期数", str(len(periods))),
        (
            "期间范围",
            "、".join(p["label"] for p in periods) if periods else "（无数据）",
        ),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("科目口径", "CAS simplified v2（L1）"),
        ("比率口径", "动态计算，不落库；见「财务比率」sheet"),
        (
            "百分比约定",
            "财务比率中 unit=% 的数值已 ×100 写出（12.5=12.5%），勿再按 Excel 百分比格式二次换算",
        ),
        ("生成", "Web_Financial_Analyse 导出"),
    ]
    ws.append(["项", "内容"])
    _style_header(ws, 2)
    for k, v in lines:
        ws.append([k, v])
    _autosize(ws, max_width=60)
    if not periods:
        ws.append([])
        ws.append(["提示", "当前筛选下无报表期间，仅导出空表结构。"])


def build_export_workbook(
    db: Session,
    company_id: int,
    *,
    period_type: str = PERIOD_ANNUAL,
    years: list[int] | None = None,
) -> tuple[bytes, str]:
    """返回 (xlsx_bytes, filename)。"""
    company = get_company(db, company_id)
    if period_type not in (PERIOD_ANNUAL, PERIOD_QUARTERLY):
        raise ValidationError("period_type 须为 annual 或 quarterly")

    bs_rows = list_balance_sheets(db, company_id)
    is_rows = list_income_statements(db, company_id)
    cf_rows = list_cash_flow_statements(db, company_id)

    periods = _collect_period_axis([bs_rows, is_rows, cf_rows], period_type, years)

    wb = Workbook()
    # 去掉默认 sheet，由说明页接管
    default = wb.active
    wb.remove(default)

    _write_readme(wb, company.name, company.code, period_type, periods)

    _write_statement_sheet(
        wb,
        "资产负债表",
        BALANCE_SHEET_GROUPS,
        periods,
        _index_by_period(bs_rows, period_type),
    )
    _write_statement_sheet(
        wb,
        "利润表",
        INCOME_STATEMENT_GROUPS,
        periods,
        _index_by_period(is_rows, period_type),
    )
    _write_statement_sheet(
        wb,
        "现金流量表",
        CASH_FLOW_GROUPS,
        periods,
        _index_by_period(cf_rows, period_type),
    )
    _write_ratio_sheet(wb, db, company_id, periods)

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    if periods:
        y0, y1 = periods[0]["year"], periods[-1]["year"]
        span = f"{y0}" if y0 == y1 else f"{y0}-{y1}"
    else:
        span = "empty"
    fname = f"{_safe_filename(company.name)}_{period_type}_{span}.xlsx"
    return data, fname
