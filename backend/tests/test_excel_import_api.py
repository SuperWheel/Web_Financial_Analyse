"""Excel 模板下载与三表导入测试。"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook


def _create_company(client, name: str = "Excel导入公司", code: str = "XLS1") -> int:
    r = client.post("/api/companies", json={"name": name, "code": code})
    assert r.status_code == 201, r.text
    return r.json()["id"]


def _build_minimal_workbook() -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # BS
    ws = wb.create_sheet("资产负债表")
    ws.append(["科目代码", "科目", "2024 年报", "2025 年报"])
    ws.append(["monetary_funds", "货币资金", 100.0, 150.0])
    ws.append(["total_assets", "资产总计", 500.0, 600.0])
    ws.append(["total_liabilities", "负债合计", 200.0, 240.0])
    ws.append(["total_equity", "所有者权益合计", 300.0, 360.0])
    ws.append(["total_current_assets", "流动资产合计", 200.0, 250.0])
    ws.append(["total_current_liabilities", "流动负债合计", 100.0, 100.0])

    # IS
    ws = wb.create_sheet("利润表")
    ws.append(["科目代码", "科目", "2024 年报", "2025 年报"])
    ws.append(["operating_revenue", "营业收入", 1000.0, 1200.0])
    ws.append(["operating_cost", "营业成本", 600.0, 700.0])
    ws.append(["net_profit", "净利润", 120.0, 150.0])

    # CF
    ws = wb.create_sheet("现金流量表")
    ws.append(["科目代码", "科目", "2024 年报", "2025 年报"])
    ws.append(["net_cash_flow_operating", "经营活动现金流量净额", 150.0, 180.0])

    # 比率应被忽略
    ws = wb.create_sheet("财务比率")
    ws.append(["分组", "比率", "单位", "公式", "2024 年报"])
    ws.append(["偿债能力", "流动比率", "倍", "x", 999])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_download_template(client):
    r = client.get(
        "/api/excel/template.xlsx",
        params={"period_type": "annual", "years": "2023,2024"},
    )
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = load_workbook(BytesIO(r.content))
    assert "资产负债表" in wb.sheetnames
    assert "利润表" in wb.sheetnames
    assert "现金流量表" in wb.sheetnames
    headers = [c.value for c in wb["资产负债表"][1]]
    assert headers[:2] == ["科目代码", "科目"]
    assert "2023 年报" in headers
    assert "2024 年报" in headers


def test_preview_and_import_create(client):
    cid = _create_company(client)
    content = _build_minimal_workbook()

    prev = client.post(
        f"/api/companies/{cid}/excel/preview",
        files={
            "file": (
                "t.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert prev.status_code == 200, prev.text
    body = prev.json()
    assert body["company_id"] == cid
    assert body["period_type"] == "annual"
    assert len(body["will_create"]) >= 3  # 至少三表×有数据期
    assert body["will_update"] == []

    imp = client.post(
        f"/api/companies/{cid}/excel/import",
        data={"overwrite": "true"},
        files={
            "file": (
                "t.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imp.status_code == 200, imp.text
    result = imp.json()
    assert len(result["created"]) >= 3
    assert result["updated"] == []

    # 校验入库金额
    bs = client.get(
        f"/api/companies/{cid}/balance-sheets",
        params={"year": 2025, "period_type": "annual"},
    )
    assert bs.status_code == 200
    rows = bs.json()
    assert len(rows) == 1
    assert rows[0]["monetary_funds"] == 150.0
    assert rows[0]["total_assets"] == 600.0

    inc = client.get(
        f"/api/companies/{cid}/income-statements",
        params={"year": 2024, "period_type": "annual"},
    )
    assert inc.json()[0]["operating_revenue"] == 1000.0


def test_import_overwrite(client):
    cid = _create_company(client, name="覆盖导入", code="XLS2")
    content = _build_minimal_workbook()
    assert (
        client.post(
            f"/api/companies/{cid}/excel/import",
            data={"overwrite": "true"},
            files={"file": ("t.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ).status_code
        == 200
    )

    # 改 2025 货币资金
    wb = load_workbook(BytesIO(content))
    ws = wb["资产负债表"]
    # 找 monetary_funds 行，改 2025 列
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "monetary_funds":
            row[3].value = 999.0
            break
    buf = BytesIO()
    wb.save(buf)
    new_content = buf.getvalue()

    r = client.post(
        f"/api/companies/{cid}/excel/import",
        data={"overwrite": "true"},
        files={
            "file": (
                "t2.xlsx",
                new_content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 200, r.text
    assert any("资产负债表/2025" in x for x in r.json()["updated"])

    bs = client.get(
        f"/api/companies/{cid}/balance-sheets",
        params={"year": 2025, "period_type": "annual"},
    )
    assert bs.json()[0]["monetary_funds"] == 999.0


def test_import_skip_when_no_overwrite(client):
    cid = _create_company(client, name="跳过存在", code="XLS3")
    content = _build_minimal_workbook()
    assert (
        client.post(
            f"/api/companies/{cid}/excel/import",
            data={"overwrite": "true"},
            files={"file": ("t.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ).status_code
        == 200
    )
    r = client.post(
        f"/api/companies/{cid}/excel/import",
        data={"overwrite": "false"},
        files={"file": ("t.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["created"] == []
    assert any("overwrite=false" in s for s in body["skipped"])


def test_export_roundtrip_import(client):
    """导出工作簿可再导入（比率 sheet 忽略）。"""
    cid = _create_company(client, name="往返", code="XLS4")
    # 先 seed 再 export
    assert (
        client.post(
            f"/api/companies/{cid}/balance-sheets",
            json={
                "year": 2024,
                "period_type": "annual",
                "monetary_funds": 42.0,
                "total_assets": 100.0,
            },
        ).status_code
        == 201
    )
    assert (
        client.post(
            f"/api/companies/{cid}/income-statements",
            json={
                "year": 2024,
                "period_type": "annual",
                "operating_revenue": 200.0,
                "net_profit": 20.0,
            },
        ).status_code
        == 201
    )
    exp = client.get(
        f"/api/companies/{cid}/export.xlsx",
        params={"period_type": "annual", "years": "2024"},
    )
    assert exp.status_code == 200

    # 新公司导入导出文件
    cid2 = _create_company(client, name="往返目标", code="XLS5")
    r = client.post(
        f"/api/companies/{cid2}/excel/import",
        data={"overwrite": "true"},
        files={
            "file": (
                "exp.xlsx",
                exp.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 200, r.text
    assert r.json()["created"]

    bs = client.get(
        f"/api/companies/{cid2}/balance-sheets",
        params={"year": 2024, "period_type": "annual"},
    )
    assert bs.json()[0]["monetary_funds"] == 42.0
    inc = client.get(
        f"/api/companies/{cid2}/income-statements",
        params={"year": 2024, "period_type": "annual"},
    )
    assert inc.json()[0]["operating_revenue"] == 200.0


def test_import_missing_company(client):
    content = _build_minimal_workbook()
    r = client.post(
        "/api/companies/99999/excel/import",
        data={"overwrite": "true"},
        files={"file": ("t.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 404
