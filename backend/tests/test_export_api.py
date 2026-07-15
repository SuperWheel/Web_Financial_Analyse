"""Excel 导出 API 测试（含财务比率 sheet）。"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def _create_company(client, name: str = "导出测试公司", code: str = "EXP1") -> int:
    r = client.post("/api/companies", json={"name": name, "code": code})
    assert r.status_code == 201, r.text
    return r.json()["id"]


def _seed(client, cid: int) -> None:
    for year, assets, cash, rev, cost, equity, liab in (
        (2024, 500.0, 100.0, 1000.0, 600.0, 300.0, 200.0),
        (2025, 600.0, 150.0, 1200.0, 700.0, 360.0, 240.0),
    ):
        assert (
            client.post(
                f"/api/companies/{cid}/balance-sheets",
                json={
                    "year": year,
                    "period_type": "annual",
                    "monetary_funds": cash,
                    "inventories": 20.0,
                    "total_current_assets": cash + 100,
                    "total_current_liabilities": 100.0,
                    "total_assets": assets,
                    "total_liabilities": liab,
                    "total_equity": equity,
                    "total_equity_parent": equity,
                },
            ).status_code
            == 201
        )
        assert (
            client.post(
                f"/api/companies/{cid}/income-statements",
                json={
                    "year": year,
                    "period_type": "annual",
                    "operating_revenue": rev,
                    "operating_cost": cost,
                    "operating_profit": rev - cost - 50,
                    "net_profit": rev - cost - 100,
                    "net_profit_parent": rev - cost - 100,
                },
            ).status_code
            == 201
        )
        assert (
            client.post(
                f"/api/companies/{cid}/cash-flow-statements",
                json={
                    "year": year,
                    "period_type": "annual",
                    "net_cash_flow_operating": 150.0 if year == 2024 else 180.0,
                },
            ).status_code
            == 201
        )


def test_export_xlsx_contains_statements_and_ratios(client):
    cid = _create_company(client)
    _seed(client, cid)

    resp = client.get(
        f"/api/companies/{cid}/export.xlsx",
        params={"period_type": "annual"},
    )
    assert resp.status_code == 200, resp.text
    assert (
        "spreadsheetml"
        in resp.headers.get("content-type", "")
    )
    assert "attachment" in resp.headers.get("content-disposition", "")

    wb = load_workbook(BytesIO(resp.content))
    assert "说明" in wb.sheetnames
    assert "资产负债表" in wb.sheetnames
    assert "利润表" in wb.sheetnames
    assert "现金流量表" in wb.sheetnames
    assert "财务比率" in wb.sheetnames

    bs = wb["资产负债表"]
    # 表头：科目代码 | 科目 | 2024 年报 | 2025 年报
    headers = [c.value for c in bs[1]]
    assert headers[0] == "科目代码"
    assert "2024 年报" in headers
    assert "2025 年报" in headers

    # 找货币资金行
    cash_row = None
    for row in bs.iter_rows(min_row=2, values_only=True):
        if row[0] == "monetary_funds":
            cash_row = row
            break
    assert cash_row is not None
    # 科目代码, 科目, 2024, 2025
    assert cash_row[2] == 100.0
    assert cash_row[3] == 150.0

    ratios = wb["财务比率"]
    r_headers = [c.value for c in ratios[1]]
    assert r_headers[:4] == ["分组", "比率", "单位", "公式"]
    assert "2024 年报" in r_headers
    assert "2025 年报" in r_headers

    # 流动比率：2024 (100+100)/100=2.0，2025 (150+100)/100=2.5
    current = None
    debt = None
    for row in ratios.iter_rows(min_row=2, values_only=True):
        if row[1] == "流动比率":
            current = row
        if row[1] == "资产负债率":
            debt = row
    assert current is not None
    assert current[2] == "倍"
    assert abs(float(current[4]) - 2.0) < 1e-6
    assert abs(float(current[5]) - 2.5) < 1e-6

    # 资产负债率 percent 导出为 ×100：2024 200/500=40
    assert debt is not None
    assert debt[2] == "%"
    assert abs(float(debt[4]) - 40.0) < 1e-4


def test_export_years_filter(client):
    cid = _create_company(client, name="过滤导出", code="EXP2")
    _seed(client, cid)
    resp = client.get(
        f"/api/companies/{cid}/export.xlsx",
        params={"period_type": "annual", "years": "2025"},
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    headers = [c.value for c in wb["资产负债表"][1]]
    assert "2025 年报" in headers
    assert "2024 年报" not in headers


def test_export_missing_company(client):
    r = client.get("/api/companies/99999/export.xlsx")
    assert r.status_code == 404


def test_export_empty_company_still_workbook(client):
    cid = _create_company(client, name="空公司", code="EXP3")
    resp = client.get(f"/api/companies/{cid}/export.xlsx")
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    assert "财务比率" in wb.sheetnames
    # 仅表头，无期间列
    headers = [c.value for c in wb["财务比率"][1]]
    assert headers == ["分组", "比率", "单位", "公式"]
